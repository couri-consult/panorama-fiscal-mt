"""Manual XLSX loader: items that don't have an API yet (beneficiometro, ranking CLP, agenda, PPP textual data).

Reads each sheet into a list[dict], same shape as the previous panorama_fiscal_df.xlsx flow.
"""
import openpyxl


def read_sheet(wb, name):
    """Read a worksheet into a list of dicts using row 1 as headers. Stops at first all-empty row."""
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    headers = None
    out = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c).strip() if c is not None else None for c in row]
            continue
        if not any(c is not None and str(c).strip() != "" for c in row):
            break
        rec = {}
        for h, v in zip(headers, row):
            if h is None:
                continue
            rec[h] = v
        out.append(rec)
    return out


def load_manual(xlsx_path):
    """Load all known manual sheets from the workbook. Missing sheets return [] (so this stays
    forward-compatible if you add/remove sheets in the file)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    return {
        "beneficios": read_sheet(wb, "beneficios"),
        "investimentos_uf": read_sheet(wb, "investimentos"),
        "clp_ranking": read_sheet(wb, "clp_ranking"),
        "ppps": read_sheet(wb, "ppps"),
        "ppps_projecao": read_sheet(wb, "ppps_projecao"),
        "agenda": read_sheet(wb, "agenda"),
    }
